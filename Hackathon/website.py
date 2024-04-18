import requests
import smtplib
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
import streamlit as st
import time

def main():
    st.title("Website Change Monitor")

    st.sidebar.header("Email Configuration")
    sender_email = st.sidebar.text_input("Sender Email")
    password = st.sidebar.text_input("Password", type="password")
    receiver_email = st.sidebar.text_input("Receiver Email")

    websites_file = st.file_uploader("Upload Excel file containing websites", type=["xlsx"])

    if sender_email and password and receiver_email and websites_file:
        try:
            workbook = load_workbook(websites_file)
            worksheet = workbook["Websites"]
            websites = [row[0] for row in worksheet.iter_rows(values_only=True)]

            previous_content = {}

            smtp_server = 'smtp.gmail.com'
            smtp_port = 587

            st.text("Monitoring websites...")

            while True:
                check_websites(sender_email, password, receiver_email, websites, previous_content, smtp_server, smtp_port)
                time.sleep(300)  # Check every 5 minutes
        except Exception as e:
            st.error(f"Error occurred: {e}")

def send_email(sender_email, password, receiver_email, website):
    subject = f'Website {website} has changed!'
    body = f'The content of {website} has changed.'
    message = MIMEText(body)
    message['Subject'] = subject
    message['From'] = sender_email
    message['To'] = receiver_email
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message.as_string())
        st.write(f"Email notification sent for {website}.")
    except Exception as e:
        st.error(f"Failed to send email notification for {website}. Error: {e}")
    finally:
        server.quit()

def check_websites(sender_email, password, receiver_email, websites, previous_content, smtp_server, smtp_port):
    for website in websites:
        try:
            response = requests.get(website)
            content = response.text

            if website not in previous_content:
                previous_content[website] = content
            elif previous_content[website] != content:
                send_email(sender_email, password, receiver_email, website)
                previous_content[website] = content
        except Exception as e:
            st.error(f"Error occurred while checking {website}: {e}")

if __name__ == "__main__":
    main()
